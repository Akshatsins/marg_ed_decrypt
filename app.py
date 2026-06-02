"""
Marg EDE - Corporate Collaboration App
Streamlit Web Interface — Modern UI

Run with:
    streamlit run app.py

Requirements:
    pip install streamlit pycryptodome openpyxl requests
"""

import base64, io, json, zlib, datetime
from Crypto.Cipher import AES
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
import streamlit as st

# ─── Page config ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Marg EDE Sync",
    page_icon="📊",
    layout="centered",
)

# ─── CSS ─────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Background */
.stApp {
    background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%);
    min-height: 100vh;
}

.block-container {
    padding: 2rem 3rem 3rem 3rem;
    max-width: 1100px;
}

/* Hero banner */
.hero {
    background: linear-gradient(135deg, #1e40af 0%, #3b82f6 50%, #06b6d4 100%);
    border-radius: 20px;
    padding: 3rem 3rem 2.5rem 3rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    box-shadow: 0 25px 50px rgba(59,130,246,0.3);
}
.hero::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: rgba(255,255,255,0.05);
    border-radius: 50%;
}
.hero::after {
    content: '';
    position: absolute;
    bottom: -30%;
    right: 15%;
    width: 200px;
    height: 200px;
    background: rgba(255,255,255,0.05);
    border-radius: 50%;
}
.hero-badge {
    display: inline-block;
    background: rgba(255,255,255,0.2);
    color: white;
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    padding: 4px 14px;
    border-radius: 20px;
    margin-bottom: 1rem;
    backdrop-filter: blur(10px);
}
.hero-title {
    font-size: 2.6rem;
    font-weight: 700;
    color: white;
    margin: 0 0 0.5rem 0;
    line-height: 1.2;
}
.hero-sub {
    font-size: 1rem;
    color: rgba(255,255,255,0.8);
    margin: 0;
    font-weight: 400;
}
.hero-stats {
    display: flex;
    gap: 2rem;
    margin-top: 2rem;
}
.hero-stat {
    text-align: center;
}
.hero-stat-num {
    font-size: 1.6rem;
    font-weight: 700;
    color: white;
}
.hero-stat-label {
    font-size: 0.75rem;
    color: rgba(255,255,255,0.65);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* Cards */
.card {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 16px;
    padding: 1.5rem;
    backdrop-filter: blur(10px);
    margin-bottom: 1.2rem;
}
.card-title {
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #94a3b8;
    margin-bottom: 1rem;
}

/* Step badges */
.step-row {
    display: flex;
    align-items: flex-start;
    gap: 1rem;
    margin-bottom: 1.2rem;
}
.step-badge {
    min-width: 32px;
    height: 32px;
    background: linear-gradient(135deg, #3b82f6, #06b6d4);
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 13px;
    font-weight: 700;
    color: white;
    flex-shrink: 0;
}
.step-content-title {
    font-size: 0.9rem;
    font-weight: 600;
    color: #f1f5f9;
    margin-bottom: 2px;
}
.step-content-desc {
    font-size: 0.8rem;
    color: #94a3b8;
}

/* Metric cards */
.metric-card {
    background: linear-gradient(135deg, rgba(59,130,246,0.15), rgba(6,182,212,0.1));
    border: 1px solid rgba(59,130,246,0.3);
    border-radius: 12px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.metric-value {
    font-size: 1.8rem;
    font-weight: 700;
    color: #60a5fa;
    line-height: 1;
}
.metric-label {
    font-size: 0.72rem;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-top: 4px;
}

/* Success / Error banners */
.success-banner {
    background: linear-gradient(135deg, rgba(16,185,129,0.2), rgba(5,150,105,0.1));
    border: 1px solid rgba(16,185,129,0.4);
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    margin: 1rem 0;
    display: flex;
    align-items: center;
    gap: 0.8rem;
}
.success-icon { font-size: 1.5rem; }
.success-text { color: #6ee7b7; font-size: 0.95rem; font-weight: 500; }

.error-banner {
    background: linear-gradient(135deg, rgba(239,68,68,0.2), rgba(185,28,28,0.1));
    border: 1px solid rgba(239,68,68,0.4);
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    margin: 1rem 0;
}
.error-title { color: #fca5a5; font-size: 0.95rem; font-weight: 600; margin-bottom: 4px; }
.error-desc  { color: #f87171; font-size: 0.85rem; font-family: monospace; }

/* Buttons */
.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #2563eb, #0891b2) !important;
    color: white !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    padding: 0.75rem 1.5rem !important;
    border-radius: 10px !important;
    border: none !important;
    letter-spacing: 0.3px;
    transition: all 0.2s ease;
    box-shadow: 0 4px 15px rgba(37,99,235,0.4) !important;
}
.stButton > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 6px 20px rgba(37,99,235,0.5) !important;
}

.stDownloadButton > button {
    width: 100%;
    background: linear-gradient(135deg, #059669, #0891b2) !important;
    color: white !important;
    font-size: 16px !important;
    font-weight: 700 !important;
    padding: 0.9rem 1.5rem !important;
    border-radius: 12px !important;
    border: none !important;
    letter-spacing: 0.3px;
    box-shadow: 0 4px 20px rgba(5,150,105,0.4) !important;
    transition: all 0.2s ease;
}
.stDownloadButton > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(5,150,105,0.5) !important;
}

/* Inputs */
.stTextInput > div > div > input {
    background: rgba(255,255,255,0.07) !important;
    border: 1px solid rgba(255,255,255,0.15) !important;
    border-radius: 8px !important;
    color: #f1f5f9 !important;
    font-size: 13px !important;
}
.stTextInput > div > div > input:focus {
    border-color: #3b82f6 !important;
    box-shadow: 0 0 0 2px rgba(59,130,246,0.2) !important;
}
label { color: #94a3b8 !important; font-size: 12px !important; font-weight: 500 !important; }

/* Expander */
.streamlit-expanderHeader {
    background: rgba(255,255,255,0.05) !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 10px !important;
    color: #cbd5e1 !important;
    font-size: 13px !important;
    font-weight: 500 !important;
}
.streamlit-expanderContent {
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    border-top: none !important;
    border-radius: 0 0 10px 10px !important;
    padding: 1rem !important;
}

/* Divider */
hr { border-color: rgba(255,255,255,0.08) !important; margin: 1.5rem 0 !important; }

/* Spinner */
.stSpinner > div { border-top-color: #3b82f6 !important; }

/* Footer */
.footer {
    text-align: center;
    color: #475569;
    font-size: 12px;
    margin-top: 2rem;
    padding-top: 1.5rem;
    border-top: 1px solid rgba(255,255,255,0.06);
}
.footer span { color: #64748b; }

/* Hide streamlit branding */
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─── Crypto helpers ───────────────────────────────────────────────────────────

def decrypt_marg(data: str, key: str) -> str:
    clean = data.replace(" ", "").replace("\n", "").replace("\r", "")
    enc = base64.b64decode(clean)
    kp = (key.encode() + b"\x00" * 16)[:16]
    cipher = AES.new(kp, AES.MODE_CBC, iv=kp)
    plain = cipher.decrypt(enc)
    return plain[:-plain[-1]].decode("utf-8")

def decompress_marg(b64: str) -> str:
    return zlib.decompress(base64.b64decode(b64), -15).decode("utf-8-sig")

def decrypt_and_parse(payload: str, key: str) -> dict:
    return json.loads(decompress_marg(decrypt_marg(payload, key)))

# ─── Excel helpers ────────────────────────────────────────────────────────────

H_FILL = PatternFill("solid", start_color="1F4E79")
H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
A_FILL = PatternFill("solid", start_color="D6E4F0")
N_FONT = Font(name="Arial", size=10)
_THIN  = Side(style="thin", color="BFBFBF")
BRD    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CTR    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT    = Alignment(horizontal="left",   vertical="center")

def _write_sheet(wb, title: str, rows: list, headers: list):
    ws = wb.create_sheet(title=title[:31])
    ws.freeze_panes = "A2"
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h.upper())
        c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, CTR, BRD
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(rows, 2):
        fill = A_FILL if ri % 2 == 0 else None
        for ci, h in enumerate(headers, 1):
            c = ws.cell(ri, ci, row.get(h, ""))
            c.font, c.alignment, c.border = N_FONT, LFT, BRD
            if fill:
                c.fill = fill
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 42)

def json_to_excel_bytes(data: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    written, summary = 0, {}

    if "Details" in data and isinstance(data["Details"], dict):
        for k, rows in data["Details"].items():
            if not isinstance(rows, list) or not rows:
                continue
            headers = list(dict.fromkeys(col for r in rows if isinstance(r, dict) for col in r.keys()))
            if not headers:
                continue
            _write_sheet(wb, k, rows, headers)
            summary[k] = len(rows)
            written += 1

    TOP_KEYS = ["Details", "Masters", "MDis", "Party", "Product", "SaleType", "Stock"]
    for k in TOP_KEYS:
        rows = data.get(k)
        if not isinstance(rows, list) or not rows:
            continue
        headers = list(dict.fromkeys(col for r in rows if isinstance(r, dict) for col in r.keys()))
        if not headers:
            continue
        _write_sheet(wb, k, rows, headers)
        summary[k] = len(rows)
        written += 1

    if written == 0:
        for k, v in data.items():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                headers = list(dict.fromkeys(col for r in v for col in r.keys()))
                _write_sheet(wb, str(k)[:31], v, headers)
                summary[k] = len(v)

    if not summary:
        ws = wb.create_sheet("Raw JSON")
        ws["A1"] = json.dumps(data, indent=2)
        summary["Raw JSON"] = 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary

# ─── API ─────────────────────────────────────────────────────────────────────

DEFAULT_URL     = "https://corporate.margerp.com/api/eOnlineData/MargCorporateEDE"
DEFAULT_COMPANY = "RAKESHCORPORATE3"
DEFAULT_MARGKEY = "1DVQEFIU2CL4RBIILE9VDA8QCK2X2BSNX3CV"
DEFAULT_COMPID  = "8820"
DEFAULT_DECKEY  = "BXOW2BIS1IGD"

def fetch_and_decrypt(company, margkey, companyid, deckey, dt_filter, index):
    resp = requests.post(
        DEFAULT_URL,
        json={"CompanyCode": company, "Datetime": dt_filter, "MargKey": margkey,
              "Index": index, "CompanyID": companyid, "APIType": "2"},
        timeout=60,
    )
    resp.raise_for_status()
    raw = resp.json()

    if isinstance(raw, str):
        encrypted = raw
    elif isinstance(raw, dict):
        encrypted = (raw.get("Data") or raw.get("data") or raw.get("Result")
                     or raw.get("result") or raw.get("EData") or raw.get("edata"))
    else:
        encrypted = raw

    if not isinstance(encrypted, str):
        raise ValueError(f"Unexpected API response: {str(raw)[:200]}")

    return json_to_excel_bytes(decrypt_and_parse(encrypted, deckey))

# ─── UI ──────────────────────────────────────────────────────────────────────

# Hero
st.markdown("""
<div class="hero">
    <div class="hero-badge">📊 Marg ERP 9+</div>
    <h1 class="hero-title">EDE Data Sync</h1>
    <p class="hero-sub">Corporate Collaboration API — fetch, decrypt & export to Excel in one click</p>
    <div class="hero-stats">
        <div class="hero-stat">
            <div class="hero-stat-num">AES‑128</div>
            <div class="hero-stat-label">Encryption</div>
        </div>
        <div class="hero-stat">
            <div class="hero-stat-num">CBC</div>
            <div class="hero-stat-label">Mode</div>
        </div>
        <div class="hero-stat">
            <div class="hero-stat-num">.xlsx</div>
            <div class="hero-stat-label">Output</div>
        </div>
        <div class="hero-stat">
            <div class="hero-stat-num">Live</div>
            <div class="hero-stat-label">API</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# Settings
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">⚙️ API Configuration</div>', unsafe_allow_html=True)

with st.expander("Show / Edit Credentials", expanded=False):
    col1, col2 = st.columns(2)
    with col1:
        company   = st.text_input("Company Code",   value=DEFAULT_COMPANY)
        companyid = st.text_input("Company ID",     value=DEFAULT_COMPID)
        deckey    = st.text_input("Decryption Key", value=DEFAULT_DECKEY, type="password")
    with col2:
        margkey   = st.text_input("Marg Key",       value=DEFAULT_MARGKEY, type="password")
        index     = st.text_input("Index",          value="0")
        dt_filter = st.text_input(
            "Date Filter",
            value="",
            placeholder="YYYY-MM-DD HH:MM:SS  (blank = ALL data)",
        )
st.markdown('</div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Fetch button
if st.button("🔄  Fetch & Decrypt Data"):
    with st.spinner("Connecting to Marg API…"):
        try:
            excel_bytes, summary = fetch_and_decrypt(
                company, margkey, companyid, deckey, dt_filter, index
            )
            st.session_state["excel_bytes"] = excel_bytes
            st.session_state["summary"]     = summary
            st.session_state["filename"]    = f"marg_ede_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
            st.session_state["error"]       = None
        except Exception as e:
            st.session_state["excel_bytes"] = None
            st.session_state["error"]       = str(e)

# Error
if st.session_state.get("error"):
    st.markdown(f"""
    <div class="error-banner">
        <div class="error-title">❌ Something went wrong</div>
        <div class="error-desc">{st.session_state['error']}</div>
    </div>
    """, unsafe_allow_html=True)

# Success + download
if st.session_state.get("excel_bytes"):
    summary  = st.session_state["summary"]
    filename = st.session_state["filename"]
    total    = sum(summary.values())

    st.markdown(f"""
    <div class="success-banner">
        <span class="success-icon">✅</span>
        <span class="success-text">Data fetched! <strong>{len(summary)}</strong> sheets · <strong>{total}</strong> total records</span>
    </div>
    """, unsafe_allow_html=True)

    # Metric cards
    cols = st.columns(min(len(summary), 4))
    for i, (sheet, count) in enumerate(summary.items()):
        with cols[i % 4]:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{count}</div>
                <div class="metric-label">{sheet}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    st.download_button(
        label="⬇️  Download Excel File",
        data=st.session_state["excel_bytes"],
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown(f"<p style='color:#475569;font-size:11px;text-align:center;margin-top:6px'>{filename}</p>",
                unsafe_allow_html=True)

# Footer
st.markdown("""
<div class="footer">
    <span>Marg ERP 9+</span> &nbsp;·&nbsp;
    <span>EDE Corporate Collaboration API</span> &nbsp;·&nbsp;
    <span>AES-128 CBC + Deflate</span>
</div>
""", unsafe_allow_html=True)
